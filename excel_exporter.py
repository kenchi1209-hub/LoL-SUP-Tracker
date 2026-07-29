import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from champion_map import champion_to_ja
from queue_map import queue_id_to_name
from summary_exporter import (
    filter_support_rows,
    group_by_champion,
    group_by_queue,
    group_by_role,
    load_my_matches,
    summarize_overall,
    to_float,
    to_int,
)
from report_exporter import role_to_name, win_to_wl

MY_MATCHES_CSV_PATH = "data/csv/my_matches.csv"
REVIEW_CSV_PATH = "data/csv/review.csv"
EXCEL_PATH = "data/excel/lol_report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F3864")
LABEL_FONT = Font(bold=True)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

WIN_FILL = PatternFill("solid", fgColor="E2EFDA")
LOSE_FILL = PatternFill("solid", fgColor="FCE4E4")

# 集計シート共通のカラム定義: (見出し, 集計dictのキー, 表示形式)
GROUP_COLUMNS = [
    ("試合数", "total", "0"),
    ("勝", "wins", "0"),
    ("敗", "losses", "0"),
    ("勝率", "win_rate", "0.0%"),
    ("平均K", "avg_kills", "0.0"),
    ("平均D", "avg_deaths", "0.0"),
    ("平均A", "avg_assists", "0.0"),
    ("KDA", "kda_ratio", None),
    ("平均CS/m", "avg_cs_per_min", "0.0"),
    ("平均VS", "avg_vision_score", "0.0"),
    ("VS/m", "avg_vision_score_per_min", "0.00"),
]

# 試合履歴シート: (見出し, my_matches.csvのカラム, 変換関数, 表示形式)
MATCH_COLUMNS = [
    ("日時", "date", None, None),
    ("キュー", "queue_id", queue_id_to_name, None),
    ("勝敗", "win", win_to_wl, None),
    ("ロール", "role", role_to_name, None),
    ("チャンピオン", "champion", champion_to_ja, None),
    ("K", "kills", to_int, "0"),
    ("D", "deaths", to_int, "0"),
    ("A", "assists", to_int, "0"),
    ("チームK", "team_kills", to_int, "0"),
    ("チームD", "team_deaths", to_int, "0"),
    ("チームA", "team_assists", to_int, "0"),
    ("CS", "cs", to_int, "0"),
    ("CS/m", "cs_per_min", to_float, "0.0"),
    ("VS", "vision_score", to_int, "0"),
    ("VS/m", "vision_score_per_min", to_float, "0.00"),
    ("ward設置", "wards_placed", to_int, "0"),
    ("ward破壊", "wards_killed", to_int, "0"),
    ("コントロールward", "control_wards_bought", to_int, "0"),
    ("ゴールド", "gold_earned", to_int, "#,##0"),
    ("与ダメージ", "total_damage_to_champions", to_int, "#,##0"),
    ("試合時間(分)", "game_duration_min", to_float, "0.0"),
    ("match_id", "match_id", None, None),
]

REVIEW_COLUMNS = [
    ("日時", "date"),
    ("キュー", "queue"),
    ("勝敗", "win"),
    ("ロール", "role"),
    ("チャンピオン", "champion"),
    ("K", "kills"),
    ("D", "deaths"),
    ("A", "assists"),
    ("CS", "cs"),
    ("VS", "vision_score"),
    ("試合状況", "game_flow"),
    ("MEMO", "memo"),
    ("GOOD", "good_point"),
    ("BAD", "bad_point"),
    ("次回テーマ", "next_theme"),
    ("match_id", "match_id"),
]


def display_width(text):
    """日本語を2文字幅として数え、列幅の目安を返す"""
    return sum(2 if ord(char) > 0x1100 else 1 for char in str(text))


def auto_fit_columns(worksheet, header_row=1, max_width=42):
    for column_cells in worksheet.columns:
        widths = [
            display_width(cell.value)
            for cell in column_cells
            if cell.row >= header_row and cell.value is not None
        ]
        if not widths:
            continue

        letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[letter].width = min(max(widths) + 2, max_width)


def write_header(worksheet, headers, row=1):
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def add_win_rate_rule(worksheet, column_index, last_row, first_row=2):
    """勝率列に50%以上=緑 / 50%未満=赤の色分けを付ける"""
    if last_row < first_row:
        return

    letter = get_column_letter(column_index)
    cell_range = f"{letter}{first_row}:{letter}{last_row}"

    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=WIN_FILL),
    )
    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0.5"], fill=LOSE_FILL),
    )


def write_summary_block(worksheet, start_row, title, summary):
    """全体成績・SUP成績をラベル/値の2列で書き出し、次に使える行番号を返す"""
    cell = worksheet.cell(row=start_row, column=1, value=title)
    cell.font = TITLE_FONT

    items = [
        ("対象試合数", summary["total"], "0"),
        ("勝ち", summary["wins"], "0"),
        ("負け", summary["losses"], "0"),
        ("勝率", summary["win_rate"], "0.0%"),
        ("平均キル", summary["avg_kills"], "0.0"),
        ("平均デス", summary["avg_deaths"], "0.0"),
        ("平均アシスト", summary["avg_assists"], "0.0"),
        ("KDA Ratio", summary["kda_ratio"], None),
        ("平均CS", summary["avg_cs"], "0.0"),
        ("平均CS/m", summary["avg_cs_per_min"], "0.0"),
        ("平均VS", summary["avg_vision_score"], "0.0"),
        ("平均VS/m", summary["avg_vision_score_per_min"], "0.00"),
    ]

    row = start_row + 1

    for label, value, number_format in items:
        label_cell = worksheet.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.border = THIN_BORDER

        value_cell = worksheet.cell(row=row, column=2, value=value)
        value_cell.border = THIN_BORDER

        if number_format:
            value_cell.number_format = number_format

        row += 1

    return row + 1


def build_summary_sheet(worksheet, rows, support_rows):
    title_cell = worksheet.cell(row=1, column=1, value="LoL SUP Gold Project レポート")
    title_cell.font = Font(size=16, bold=True, color="1F3864")

    dates = [r["date"] for r in rows if r.get("date")]
    period = f"対象期間：{min(dates)} ～ {max(dates)}" if dates else "対象期間：-"

    worksheet.cell(row=2, column=1, value=period)
    worksheet.cell(
        row=3,
        column=1,
        value=f"出力日時：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )

    next_row = write_summary_block(worksheet, 5, "■ 全体成績", summarize_overall(rows))

    if support_rows:
        write_summary_block(worksheet, next_row, "■ SUP成績", summarize_overall(support_rows))

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 16


def build_group_sheet(worksheet, summaries, first_header, first_key):
    headers = [first_header] + [label for label, _, _ in GROUP_COLUMNS]
    write_header(worksheet, headers)

    for row_index, summary in enumerate(summaries, start=2):
        name_cell = worksheet.cell(row=row_index, column=1, value=summary[first_key])
        name_cell.border = THIN_BORDER

        for column_offset, (_, key, number_format) in enumerate(GROUP_COLUMNS, start=2):
            cell = worksheet.cell(row=row_index, column=column_offset, value=summary[key])
            cell.border = THIN_BORDER

            if number_format:
                cell.number_format = number_format

    last_row = len(summaries) + 1

    add_win_rate_rule(worksheet, column_index=5, last_row=last_row)
    worksheet.freeze_panes = "A2"

    if summaries:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    auto_fit_columns(worksheet)


def build_match_sheet(worksheet, rows):
    write_header(worksheet, [label for label, _, _, _ in MATCH_COLUMNS])

    for row_index, row in enumerate(rows, start=2):
        for column_index, (_, key, converter, number_format) in enumerate(
            MATCH_COLUMNS, start=1
        ):
            raw_value = row.get(key, "")
            value = converter(raw_value) if converter else raw_value

            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.border = THIN_BORDER

            if number_format:
                cell.number_format = number_format

        # 勝敗列（3列目）だけ色分けして一覧性を上げる
        result_cell = worksheet.cell(row=row_index, column=3)
        result_cell.fill = WIN_FILL if result_cell.value == "W" else LOSE_FILL
        result_cell.alignment = Alignment(horizontal="center")

    last_row = len(rows) + 1

    worksheet.freeze_panes = "B2"

    if rows:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(MATCH_COLUMNS))}{last_row}"

    auto_fit_columns(worksheet)


def build_review_sheet(worksheet, review_rows):
    write_header(worksheet, [label for label, _ in REVIEW_COLUMNS])

    for row_index, row in enumerate(review_rows, start=2):
        for column_index, (_, key) in enumerate(REVIEW_COLUMNS, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=row.get(key, ""))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=key in {
                "game_flow",
                "memo",
                "good_point",
                "bad_point",
                "next_theme",
            })

    last_row = len(review_rows) + 1

    worksheet.freeze_panes = "B2"

    if review_rows:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_COLUMNS))}{last_row}"

    auto_fit_columns(worksheet, max_width=30)


def load_review_rows(review_csv_path=REVIEW_CSV_PATH):
    import csv

    if not os.path.exists(review_csv_path):
        return []

    with open(review_csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    rows.sort(key=lambda r: r.get("date", ""), reverse=True)
    return rows


def export_excel_report(
    my_matches_csv_path=MY_MATCHES_CSV_PATH,
    review_csv_path=REVIEW_CSV_PATH,
    excel_path=EXCEL_PATH,
):
    rows = load_my_matches(my_matches_csv_path)

    if not rows:
        print("lol_report.xlsxを作成できませんでした（my_matches.csvが空です）")
        return None

    support_rows = filter_support_rows(rows)
    review_rows = load_review_rows(review_csv_path)

    workbook = Workbook()

    build_summary_sheet(workbook.active, rows, support_rows)
    workbook.active.title = "サマリー"

    build_group_sheet(workbook.create_sheet("ロール別"), group_by_role(rows), "ロール", "role")
    build_group_sheet(workbook.create_sheet("キュー別"), group_by_queue(rows), "キュー", "queue")
    build_group_sheet(
        workbook.create_sheet("チャンピオン別"), group_by_champion(rows), "チャンピオン", "champion"
    )

    if support_rows:
        build_group_sheet(
            workbook.create_sheet("SUPチャンピオン別"),
            group_by_champion(support_rows),
            "チャンピオン",
            "champion",
        )

    build_match_sheet(workbook.create_sheet("試合履歴"), rows)

    if review_rows:
        build_review_sheet(workbook.create_sheet("レビュー"), review_rows)

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    workbook.save(excel_path)

    print(f"lol_report.xlsx 出力完了: {len(rows)} 件 / {excel_path}")
    return excel_path


if __name__ == "__main__":
    export_excel_report()
